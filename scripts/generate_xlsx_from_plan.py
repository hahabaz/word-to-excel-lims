#!/usr/bin/env python3
"""Generate an Excel 2019-compatible XLSX from layout_plan.json.

Uses openpyxl and avoids drawings, controls, text boxes, and images.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit("ERROR: openpyxl is required. Install with: pip install openpyxl") from exc

THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
NO_SIDE = Side(style=None)


def load_plan(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def border_for(kind: str) -> Border:
    if kind == "none":
        return Border()
    if kind == "bottom":
        return Border(bottom=THIN)
    if kind in {"box", "grid"}:
        return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def apply_style(cell, spec: dict) -> None:
    style = spec.get("style", "normal")
    font_size = float(spec.get("font_size") or (14 if style == "title" else 10.5))
    bold = bool(spec.get("font_bold", style in {"title", "label", "table_header"}))
    cell.font = Font(name="宋体", size=font_size, bold=bold)
    align = spec.get("align", "left")
    valign = spec.get("valign", "center")
    vertical = "center" if valign == "center" else valign
    horizontal = {"left": "left", "center": "center", "right": "right"}.get(align, "left")
    cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=bool(spec.get("wrap", True)))
    cell.border = border_for(spec.get("border", "box"))
    fill_color = spec.get("fill_color")
    # Default is intentionally no fill. Some LIMS templates, including the
    # seeded A012 case, should remain completely unfilled; use fills only when
    # layout_plan.json explicitly requests one based on source/reference files.
    if fill_color:
        cell.fill = PatternFill(fill_type="solid", fgColor=str(fill_color).replace("#", ""))


def style_merged_border(ws, row: int, col: int, rowspan: int, colspan: int, kind: str) -> None:
    if kind not in {"box", "grid", "bottom"}:
        return
    end_row = row + rowspan - 1
    end_col = col + colspan - 1
    for r in range(row, end_row + 1):
        for c in range(col, end_col + 1):
            left = THIN if c == col and kind in {"box", "grid"} else NO_SIDE
            right = THIN if c == end_col and kind in {"box", "grid"} else NO_SIDE
            top = THIN if r == row and kind in {"box", "grid"} else NO_SIDE
            bottom = THIN if (r == end_row and kind in {"box", "grid", "bottom"}) else NO_SIDE
            ws.cell(r, c).border = Border(left=left, right=right, top=top, bottom=bottom)


def safe_merge(ws, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    if start_row == end_row and start_col == end_col:
        return
    if start_row > end_row or start_col > end_col:
        raise ValueError(f"Invalid merge range: {start_row},{start_col}:{end_row},{end_col}")
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)


def set_page_setup(ws, setup: dict, max_row: int, max_col: int) -> None:
    orientation = setup.get("orientation", "portrait")
    ws.page_setup.orientation = "landscape" if orientation == "landscape" else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_view.showGridLines = True
    ws.page_margins.left = float(setup.get("margins", {}).get("left", 0.35))
    ws.page_margins.right = float(setup.get("margins", {}).get("right", 0.35))
    ws.page_margins.top = float(setup.get("margins", {}).get("top", 0.45))
    ws.page_margins.bottom = float(setup.get("margins", {}).get("bottom", 0.45))
    ws.page_margins.header = float(setup.get("margins", {}).get("header", 0.2))
    ws.page_margins.footer = float(setup.get("margins", {}).get("footer", 0.2))
    ws.print_options.horizontalCentered = bool(setup.get("horizontal_centered", True))
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = int(setup.get("fit_to_width", 1) or 1)
    fit_height = setup.get("fit_to_height", None)
    ws.page_setup.fitToHeight = int(fit_height) if fit_height else 0
    last = f"{get_column_letter(max_col)}{max_row}"
    ws.print_area = setup.get("print_area") or f"A1:{last}"


def generate(plan: dict, out_path: Path) -> None:
    wb = Workbook()
    # Remove default sheet after creating planned sheets if necessary.
    default = wb.active
    for sidx, sheet_plan in enumerate(plan.get("workbook", {}).get("sheets", [])):
        ws = default if sidx == 0 else wb.create_sheet()
        ws.title = sheet_plan.get("name", f"Sheet{sidx + 1}")[:31]
        ws.sheet_view.showGridLines = True
        base = sheet_plan.get("base_grid", {})
        base_cols = int(base.get("base_columns", 32))
        width = float(base.get("equal_column_width", 2.4))
        for c in range(1, base_cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = width
        default_height = float(base.get("default_row_height", 18))
        row_cursor = 1
        for section in sheet_plan.get("sections", []):
            row_cursor += int(section.get("top_spacing", 0) or 0)
            for row_spec in section.get("rows", []):
                ws.row_dimensions[row_cursor].height = float(row_spec.get("height", default_height) or default_height)
                for cell_spec in row_spec.get("cells", []):
                    start_col = int(cell_spec.get("start_col", 1))
                    colspan = int(cell_spec.get("colspan", 1))
                    rowspan = int(cell_spec.get("rowspan", 1) or 1)
                    end_col = min(base_cols, start_col + colspan - 1)
                    end_row = row_cursor + rowspan - 1
                    safe_merge(ws, row_cursor, start_col, end_row, end_col)
                    top_left = ws.cell(row_cursor, start_col)
                    top_left.value = str(cell_spec.get("text", ""))
                    apply_style(top_left, cell_spec)
                    style_merged_border(ws, row_cursor, start_col, rowspan, end_col - start_col + 1, cell_spec.get("border", "box"))
                row_cursor += 1
        max_row = max(1, row_cursor - 1)
        set_page_setup(ws, sheet_plan.get("page_setup", {}), max_row, base_cols)
    if len(plan.get("workbook", {}).get("sheets", [])) == 0:
        wb.remove(default)
        ws = wb.create_sheet("Sheet1")
        ws.sheet_view.showGridLines = True
    wb.properties.creator = plan.get("workbook", {}).get("creator", "word-to-excel-lims skill")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate .xlsx from layout_plan.json")
    parser.add_argument("layout_plan", type=Path)
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args(argv)
    try:
        plan = load_plan(args.layout_plan)
        generate(plan, args.out)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    print(f"Wrote {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
